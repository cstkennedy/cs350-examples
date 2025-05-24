import sys
from typing import TextIO

from openpyxl import Workbook

from examples.roster import Roster
from examples.student import Student

"""
Demonstrate simple Roster and Student ADTs.
"""


def generate_roster() -> Roster:
    """
    This is a non-trivial main function.
    """

    john = Student("John")
    tom = Student("Tom")
    jay = Student("Jay")
    oscar = Student("Oscar")

    all_students = [john, tom, jay, oscar]

    cs330 = Roster(4, "CS 330")

    for stu in all_students:
        _ = cs330.enroll(stu)

    return cs330


def demo_simple_table(destination: str = "report.xlsx") -> None:
    """
    Output a 3x3 table with the cells labeled as A1, A2... C2, C3
    """
    workbook = Workbook()
    sheet = workbook.create_sheet(title="Report")

    # Output 3 Rows (A, B, C)
    for idx_row in range(1, 3):
        for idx_col in range(ord("A"), ord("D")):
            # Set up the cell data
            cell_name = f"{chr(idx_col)}{idx_row}"

            # Create the cell and write to it
            sheet[cell_name] = cell_name

    # Write and save the Excel document
    workbook.save(destination)


def demo_roster_table(roster: Roster, destination: str = "report.xlsx") -> None:
    """
    Output a table with student names in the first column and grades in the
    second column.

    TODO: replace hardcoded "4.0" grades with actual grade values.
    """

    workbook = Workbook()
    sheet = workbook.create_sheet(title="Report")

    # Output the Header
    sheet.append(["Student", "Grade"])

    # Output all students
    for stu in roster.list_enrolled_students():
        # Create and output a student row
        row = [stu.name, 4.0]
        sheet.append(row)

    # Write and save the Excel document
    workbook.save(destination)


def main() -> None:
    roster = generate_roster()

    demo_simple_table()
    demo_roster_table(roster, destination="report_students.xlsx")


if __name__ == "__main__":
    main()
