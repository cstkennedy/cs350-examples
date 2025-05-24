from openpyxl import Workbook

from examples.roster import Roster


def demo_simple_table(destination: str = "report.xlsx") -> None:
    """
    Output a 3x3 table with the cells labeled as A1, A2... C2, C3
    """
    workbook = Workbook()
    sheet = workbook.create_sheet(title="Report")

    # Output 3 Rows (A, B, C)
    for idx_row in range(1, 3):
        for idx_col in range(ord("A"), ord("D")):
            # Set up the cell data
            cell_name = f"{chr(idx_col)}{idx_row}"

            # Create the cell and write to it
            sheet[cell_name] = cell_name

    # Write and save the Excel document
    workbook.save(destination)


def demo_roster_table(roster: Roster, destination: str = "report.xlsx") -> None:
    """
    Output a table with student names in the first column and grades in the
    second column.

    TODO: replace hardcoded "4.0" grades with actual grade values.
    """

    workbook = Workbook()
    sheet = workbook.create_sheet(title="Report")

    # Output the Header
    sheet.append(["Student", "Grade"])

    # Output all students
    for stu in roster.list_enrolled_students():
        # Create and output a student row
        row = [stu.name, 4.0]
        sheet.append(row)

    # Write and save the Excel document
    workbook.save(destination)
