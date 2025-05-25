from hamcrest import assert_that, contains_exactly, equal_to
from openpyxl import Workbook, load_workbook

from driver import generate_roster
from simple_excel_writer import *


def test_demo_simple_table() -> None:
    demo_simple_table()

    workbook = load_workbook("report.xlsx")
    sheet = workbook["Report"]

    assert_that(sheet["A1"].value, equal_to("A1"))
    assert_that(sheet["A2"].value, equal_to("A2"))

    assert_that(sheet["B1"].value, equal_to("B1"))
    assert_that(sheet["B2"].value, equal_to("B2"))

    assert_that(sheet["C1"].value, equal_to("C1"))
    assert_that(sheet["C2"].value, equal_to("C2"))

    workbook.close()


def test_demo_roster_table() -> None:
    roster = generate_roster()
    demo_roster_table(roster)

    workbook = load_workbook("report_students.xlsx")
    sheet = workbook["Report"]

    assert_that(
        list(
            sheet.iter_rows(min_row=0, max_col=2, max_row=5, values_only=True)
        ),
        contains_exactly(
            ("Student", "Grade"),
            ("Tom", 4),
            ("Jay", 4),
            ("John", 4),
            ("Oscar", 4),
        ),
    )

    workbook.close()
